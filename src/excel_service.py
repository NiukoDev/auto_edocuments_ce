from __future__ import annotations

from collections import Counter
from datetime import datetime
from pathlib import Path
from shutil import copy2
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import DATA_SHEET_NAME, DOCUMENT_PATTERNS, DS_COLUMNS, DS_SHEET_NAME, MONTHLY_COLUMNS, REQUIRED_DATA_COLUMNS
from file_scanner import build_file_index, has_document
from models import AnalysisResult, DataRecord, GenerationResult


class ExcelValidationError(Exception):
    pass


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _build_key(seccion_aduanera: Any, patente: Any, pedimento: Any) -> str:
    return f"{_as_text(seccion_aduanera)}-{_as_text(patente)}-{_as_text(pedimento)}"


def _format_operation_type(value: Any) -> str:
    operation_type = _as_text(value)
    return {"1": "Importacion", "2": "Exportacion"}.get(operation_type, operation_type)


def _is_supported_excel(path: Path) -> bool:
    return path.suffix.lower() in {".xlsx", ".xlsm"}


def analyze_data_file(data_path: str | Path, month: str) -> AnalysisResult:
    path = Path(data_path)
    if not path.exists():
        raise ExcelValidationError("El archivo DATA no existe.")
    if not _is_supported_excel(path):
        raise ExcelValidationError("El archivo DATA debe ser .xlsx o .xlsm.")
    if path.name.startswith("~$"):
        raise ExcelValidationError("Seleccionaste un archivo temporal de Excel. Cierra Excel y elige el archivo real.")

    workbook = load_workbook(path, read_only=True, data_only=True)
    if DATA_SHEET_NAME not in workbook.sheetnames:
        raise ExcelValidationError(f'El DATA debe contener una hoja llamada exactamente "{DATA_SHEET_NAME}".')

    sheet = workbook[DATA_SHEET_NAME]
    headers = [_normalize_cell(cell.value) for cell in sheet[1]]
    header_positions = {name: index for index, name in enumerate(headers)}
    missing = [name for name in REQUIRED_DATA_COLUMNS if name not in header_positions]
    if missing:
        raise ExcelValidationError("Faltan columnas obligatorias: " + ", ".join(missing))

    records: list[DataRecord] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value is not None and str(value).strip() != "" for value in row):
            continue

        patente = row[header_positions["Patente"]]
        pedimento = row[header_positions["Pedimento"]]
        seccion = row[header_positions["SeccionAduanera"]]
        tipo_operacion = row[header_positions["TipoOperacion"]]
        clave_documento = row[header_positions["ClaveDocumento"]]
        fecha_pago_real = row[header_positions["FechaPagoReal"]]

        if not patente or not pedimento or not seccion:
            continue

        llave = _build_key(seccion, patente, pedimento)
        records.append(
            DataRecord(
                patente=_as_text(patente),
                pedimento=_as_text(pedimento),
                seccion_aduanera=_as_text(seccion),
                llave=llave,
                tipo_operacion=_format_operation_type(tipo_operacion),
                clave_documento=_as_text(clave_documento),
                fecha_pago_real=fecha_pago_real,
                estatus="PENDIENTE",
                mes=month,
            )
        )

    if not records:
        raise ExcelValidationError("No se encontraron registros validos en Datos Generales.")

    key_counts = Counter(record.llave for record in records)
    duplicate_keys = sorted(key for key, count in key_counts.items() if count > 1)
    return AnalysisResult(
        data_path=path,
        sheet_name=DATA_SHEET_NAME,
        total_records=len(records),
        duplicate_keys=duplicate_keys,
        records=records,
    )


def inspect_output_file(output_path: str | Path, month: str) -> tuple[bool, bool]:
    path = Path(output_path)
    if not path.exists():
        return False, False
    workbook = load_workbook(path, read_only=False, data_only=False)
    has_month_sheet = month in workbook.sheetnames
    has_month_rows = False
    if DS_SHEET_NAME in workbook.sheetnames:
        sheet = workbook[DS_SHEET_NAME]
        headers = [_normalize_cell(cell.value) for cell in sheet[1]]
        if "MES" in headers:
            month_col = headers.index("MES") + 1
            for row in range(2, sheet.max_row + 1):
                if _normalize_cell(sheet.cell(row, month_col).value).upper() == month:
                    has_month_rows = True
                    break
    return has_month_sheet, has_month_rows


def generate_result_file(
    output_path: str | Path,
    analysis: AnalysisResult,
    month: str,
    overwrite_month: bool,
    expedients_folder: str | Path | None = None,
    create_backup: bool = True,
) -> GenerationResult:
    path = Path(output_path)
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ExcelValidationError("El resultado debe guardarse como .xlsx o .xlsm.")
    if path.name.startswith("~$"):
        raise ExcelValidationError("No puedes usar un archivo temporal de Excel como resultado.")

    backup_path: Path | None = None
    month_replaced = False

    if path.exists():
        if create_backup:
            timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
            backup_path = path.with_name(f"{path.stem}.backup-{timestamp}{path.suffix}")
            copy2(path, backup_path)
        workbook = load_workbook(path)
    else:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

    _ensure_ds_sheet(workbook)

    has_month_sheet, has_month_rows = _workbook_has_month(workbook, month)
    if (has_month_sheet or has_month_rows) and not overwrite_month:
        raise ExcelValidationError(f"El mes {month} ya existe. Activa la reescritura para reemplazarlo.")

    if overwrite_month:
        month_replaced = has_month_sheet or has_month_rows
        if has_month_sheet:
            workbook.remove(workbook[month])
        _remove_month_from_ds(workbook[DS_SHEET_NAME], month)

    if month in workbook.sheetnames:
        raise ExcelValidationError(f"La hoja {month} ya existe y no se pudo reemplazar.")

    _append_to_ds(workbook[DS_SHEET_NAME], analysis.records)
    indexed_names = build_file_index(expedients_folder) if expedients_folder else []
    document_counts = _create_month_sheet(workbook, month, analysis.records, indexed_names)
    _order_sheets(workbook, month)

    try:
        workbook.save(path)
    except PermissionError as exc:
        raise ExcelValidationError("No se pudo guardar. Cierra el archivo resultado si esta abierto en Excel.") from exc

    return GenerationResult(
        output_path=path,
        backup_path=backup_path,
        records_written=len(analysis.records),
        month_replaced=month_replaced,
        generated_at=datetime.now(),
        pd_found=document_counts["PD"]["found"],
        pd_missing=document_counts["PD"]["missing"],
        ps_found=document_counts["PS"]["found"],
        ps_missing=document_counts["PS"]["missing"],
    )


def _ensure_ds_sheet(workbook: Workbook) -> None:
    if DS_SHEET_NAME not in workbook.sheetnames:
        sheet = workbook.create_sheet(DS_SHEET_NAME, 0)
        sheet.append(DS_COLUMNS)
        _style_header(sheet, 1, len(DS_COLUMNS))
        _resize_columns(sheet)
        return

    sheet = workbook[DS_SHEET_NAME]
    headers = [_normalize_cell(cell.value) for cell in sheet[1]]
    if headers[: len(DS_COLUMNS)] != DS_COLUMNS:
        if sheet.max_row == 1 and not any(headers):
            for index, header in enumerate(DS_COLUMNS, start=1):
                sheet.cell(1, index).value = header
            _style_header(sheet, 1, len(DS_COLUMNS))
            return
        raise ExcelValidationError("La hoja DS existente no tiene la estructura esperada.")


def _workbook_has_month(workbook: Workbook, month: str) -> tuple[bool, bool]:
    has_sheet = month in workbook.sheetnames
    has_rows = False
    if DS_SHEET_NAME in workbook.sheetnames:
        sheet = workbook[DS_SHEET_NAME]
        headers = [_normalize_cell(cell.value) for cell in sheet[1]]
        if "MES" in headers:
            month_col = headers.index("MES") + 1
            for row in range(2, sheet.max_row + 1):
                if _normalize_cell(sheet.cell(row, month_col).value).upper() == month:
                    has_rows = True
                    break
    return has_sheet, has_rows


def _remove_month_from_ds(sheet, month: str) -> None:
    headers = [_normalize_cell(cell.value) for cell in sheet[1]]
    month_col = headers.index("MES") + 1
    for row in range(sheet.max_row, 1, -1):
        if _normalize_cell(sheet.cell(row, month_col).value).upper() == month:
            sheet.delete_rows(row, 1)


def _append_to_ds(sheet, records: list[DataRecord]) -> None:
    for record in records:
        sheet.append(
            [
                record.patente,
                record.pedimento,
                record.seccion_aduanera,
                record.llave,
                record.tipo_operacion,
                record.clave_documento,
                record.fecha_pago_real,
                record.estatus,
                record.mes,
            ]
        )
    _style_header(sheet, 1, len(DS_COLUMNS))
    _resize_columns(sheet)


def _create_month_sheet(
    workbook: Workbook,
    month: str,
    records: list[DataRecord],
    indexed_names: list[str],
) -> dict[str, dict[str, int]]:
    sheet = workbook.create_sheet(month)
    sheet.append(MONTHLY_COLUMNS)
    _style_header(sheet, 1, len(MONTHLY_COLUMNS))

    current_group: tuple[str, str] | None = None
    document_counts = {column: {"found": 0, "missing": 0} for column in DOCUMENT_PATTERNS}
    sorted_records = sorted(records, key=lambda item: (_as_text(item.tipo_operacion), item.clave_documento, item.llave))
    for record in sorted_records:
        group = (_as_text(record.tipo_operacion), record.clave_documento)
        if group != current_group:
            current_group = group
            sheet.append([f"TipoOperacion: {group[0]} | ClaveDocumento: {group[1]}"] + [""] * (len(MONTHLY_COLUMNS) - 1))
            group_row = sheet.max_row
            sheet.merge_cells(start_row=group_row, start_column=1, end_row=group_row, end_column=len(MONTHLY_COLUMNS))
            group_cell = sheet.cell(group_row, 1)
            group_cell.fill = PatternFill("solid", fgColor="172033")
            group_cell.font = Font(color="67E8F9", bold=True)
            group_cell.alignment = Alignment(horizontal="left", vertical="center")

        document_values = {column: "" for column in DOCUMENT_PATTERNS}
        missing_documents: list[str] = []
        comments = "Pendiente de validacion documental"
        if indexed_names:
            for column, pattern in DOCUMENT_PATTERNS.items():
                if has_document(record, indexed_names, pattern):
                    document_counts[column]["found"] += 1
                    document_values[column] = "a"
                else:
                    document_counts[column]["missing"] += 1
                    document_values[column] = "x"
                    missing_documents.append(f"{column}: {pattern}")
            if missing_documents:
                comments = "Falta " + ", ".join(missing_documents)

        sheet.append(
            [
                record.llave,
                record.clave_documento,
                record.fecha_pago_real,
                document_values["PD"],
                document_values["PS"],
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "PENDIENTE",
                comments,
            ]
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    _resize_columns(sheet)
    return document_counts


def _style_header(sheet, row: int, columns: int) -> None:
    fill = PatternFill("solid", fgColor="1F2937")
    font = Font(color="FFFFFF", bold=True)
    alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, columns + 1):
        cell = sheet.cell(row, col)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment


def _resize_columns(sheet) -> None:
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells[:200]:
            value = cell.value
            if value is not None:
                max_length = max(max_length, len(str(value)))
        sheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 45)


def _order_sheets(workbook: Workbook, month: str) -> None:
    if DS_SHEET_NAME in workbook.sheetnames:
        ds_sheet = workbook[DS_SHEET_NAME]
        workbook._sheets.remove(ds_sheet)
        workbook._sheets.insert(0, ds_sheet)
